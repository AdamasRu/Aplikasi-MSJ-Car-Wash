from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from datetime import datetime

def export_excel(transaksi_manager):

    wb = Workbook()

    ws = wb.active

    ws.title = "Laporan Transaksi"

    headers = [
        "Invoice",
        "Nama",
        "No WhatsApp",
        "Plat Nomor",
        "Layanan",
        "Harga",
        "Tanggal"
    ]

    ws.append(headers)

    # =========================================
    # STYLE HEADER
    # =========================================

    fill = PatternFill(
        start_color="00FFFF",
        end_color="00FFFF",
        fill_type="solid"
    )

    thin = Side(
        border_style="thin",
        color="000000"
    )

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="000000"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

    # =========================================
    # DATA TRANSAKSI
    # =========================================

    total = 0

    for trx in transaksi_manager.daftar_transaksi:

        ws.append([

            trx["invoice"],

            trx["nama"],

            trx["no_hp"],

            trx["plat"],

            trx["layanan"],

            trx["harga"],

            trx["tanggal"]

        ])

        total += trx["harga"]

    # =========================================
    # TOTAL
    # =========================================

    ws.append([])

    ws.append([
        "",
        "",
        "",
        "",
        "TOTAL",
        total
    ])

    # =========================================
    # AUTO WIDTH
    # =========================================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[column_letter].width = adjusted_width

    # =========================================
    # SAVE FILE
    # =========================================

    filename = (
        f"Laporan_Transaksi_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb.save(filename)

    return filename